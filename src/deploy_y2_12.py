"""Step 12 — Y2 TeLI transcript scoring (deployment).

Scores Y2 classroom transcripts with the trained pipeline and writes, per
classroom-session subfolder, one multi-sheet xlsx (one sheet per child POV
file), sortable for human review.

Two-stage, budget-bounded:
  1. Bi-encoder cosine scores EVERY kept utterance (local, free) -> a
     science-similarity `cosine_score`, `top_match_in_corpus`, `predicted_subtype`.
  2. gpt-oss-120b re-ranks only the highest-cosine utterances, under a hard
     dollar cap tracked by a live cost ledger that stops before the limit.

Full LLM re-ranking of all ~1.3M utterances is intentionally NOT attempted
(~$2.6k); cosine gives full coverage and the budget is spent where science is.

Usage:
    python src/deploy_y2_12.py --folders Classroom74_110525           # pilot
    python src/deploy_y2_12.py --all                                  # all 85
    python src/deploy_y2_12.py --folders Classroom74_110525 --stub    # offline
    python src/deploy_y2_12.py --all --dollar-cap 20 --top-k 10
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import numpy as np
import pandas as pd

from src.data_loader_1 import DEFAULT_PROCESSED_DIR

DEFAULT_CONFIG_PATH = _PROJECT_ROOT / "config" / "deploy.json"
LABEL_SCIENCE = "SCIENCE_TALK"
LABEL_NOT = "NOT_SCIENCE_TALK"

TRANSCRIPT_COL = "Transcription (Confidence)"
_CONF_RE = re.compile(r"\((\d+(?:\.\d+)?)\s*%\)\s*$")
_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")


def load_config(config_path: Path = DEFAULT_CONFIG_PATH) -> dict:
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Ingestion
# ---------------------------------------------------------------------------

def parse_transcription(cell) -> tuple[str, float | None]:
    """Split 'text (57.6%)' -> ('text', 0.576). No suffix -> (text, None)."""
    s = "" if cell is None else str(cell).strip()
    m = _CONF_RE.search(s)
    if m:
        conf = float(m.group(1)) / 100.0
        text = s[: m.start()].strip()
        return text, conf
    return s, None


def child_id_from_filename(name: str) -> str:
    """'65-74-943-170_110525_...' -> '170'; suffix-less base id -> 'base'."""
    stem = Path(name).name.split("_", 1)[0]
    parts = stem.split("-")
    return parts[3] if len(parts) >= 4 else "base"


def sheet_name_for(child_id: str, used: set[str]) -> str:
    """Excel-safe, <=31 char, unique sheet name."""
    base = _INVALID_SHEET.sub("", f"child_{child_id}")[:31] or "sheet"
    name, i = base, 1
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def ingest_file(path: Path, cfg: dict) -> list[dict]:
    """Read one transcript xlsx -> kept utterance records."""
    keep = set(cfg.get("speaker_keep", ["adult", "child"]))
    conf_min = cfg.get("confidence_min", 0.0)
    child_id = child_id_from_filename(path.name)
    try:
        df = pd.read_excel(path, engine="openpyxl")
    except Exception as e:  # noqa: BLE001
        print(f"  [warn] could not read {path.name}: {e}")
        return []
    if TRANSCRIPT_COL not in df.columns:
        print(f"  [warn] {path.name} missing '{TRANSCRIPT_COL}'")
        return []

    recs: list[dict] = []
    for pos, row in enumerate(df.itertuples(index=False)):
        rowd = dict(zip(df.columns, row))
        raw = rowd.get(TRANSCRIPT_COL)
        if raw is None or str(raw).strip() in ("-", "", "nan"):
            continue
        text, conf = parse_transcription(raw)
        if not text or text == "-":
            continue
        speaker = str(rowd.get("Speaker", "")).strip()
        if keep and speaker not in keep:
            continue
        if conf is not None and conf < conf_min:
            continue
        recs.append({
            "child_id": child_id,
            "source_file": path.name,
            "line_no": pos + 2,  # +2: 1-based + header row -> spreadsheet row
            "seg_start": str(rowd.get("Segment Start Time", "")),
            "seg_end": str(rowd.get("Segment End Time", "")),
            "speaker": speaker,
            "utterance": text,
            "confidence": conf,
        })
    return recs


def _primary_subtype(val) -> str:
    if isinstance(val, (list, tuple, np.ndarray)):
        items = [str(x) for x in val if x is not None]
        return "|".join(sorted(items)) if items else "unknown"
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "unknown"
    return str(val)


# ---------------------------------------------------------------------------
# Stage 1 — cosine over the anchor bank (all utterances)
# ---------------------------------------------------------------------------

def cosine_score_all(pipe, records: list[dict], *, batch_size: int = 64,
                     verbose: bool = True) -> None:
    """Attach cosine_score / top_match / subtype to every record (in place)."""
    from src.biencoder_8 import _encode
    if not records:
        return
    # Dedupe identical utterances so we encode each text once.
    uniq = sorted({r["utterance"] for r in records})
    if verbose:
        print(f"  cosine: encoding {len(uniq)} unique utterances "
              f"({len(records)} total)...")
    emb = _encode(pipe.model, uniq, pipe.query_prefix, batch_size=batch_size)
    sims = emb @ pipe.corpus_emb.T                    # (U, D)
    best = np.argmax(sims, axis=1)
    best_score = sims[np.arange(len(uniq)), best]
    by_text = {}
    for i, t in enumerate(uniq):
        j = int(best[i])
        uid = pipe.meta_utt_ids[j]
        by_text[t] = {
            "cosine_score": float(best_score[i]),
            "top_match_utt_id": uid,
            "top_match_utterance": pipe.meta_utts[j],
            "predicted_subtype": _primary_subtype(pipe.subtype_map.get(uid)),
        }
    for r in records:
        info = by_text[r["utterance"]]
        r.update(info)
        # defaults for the (possibly skipped) LLM stage
        r.setdefault("llm_score", None)
        r.setdefault("rationale", None)
        r.setdefault("degraded", None)
        r["scored_by"] = "cosine"
        r["top_match_in_corpus"] = info["top_match_utterance"]
        r["top_match_score"] = info["cosine_score"]


# ---------------------------------------------------------------------------
# Stage 2 — budgeted LLM re-rank
# ---------------------------------------------------------------------------

def budgeted_rerank(pipe, records: list[dict], cfg: dict, *,
                    max_rerank: int | None = None,
                    dollar_budget: float | None = None,
                    reset: bool = True, verbose: bool = True) -> dict:
    """Re-rank highest-cosine utterances until the dollar cap; in place.

    `max_rerank` hard-caps the number of utterances re-ranked regardless of
    dollars (used by the offline `--stub` path to keep the smoke fast).
    `dollar_budget` overrides `cfg["dollar_cap"]` (per-folder cumulative
    allowance for the incremental runner). `reset=False` keeps the ledger
    running across folders so the cap is enforced on cumulative spend.
    """
    top_k = cfg.get("top_k", 10)
    cap = dollar_budget if dollar_budget is not None else cfg.get("dollar_cap", 15.0)
    cost = cfg.get("cost_per_call", 0.0001)
    max_calls = int(math.floor(cap / cost)) if cost > 0 else 0

    order = sorted(records, key=lambda r: r["cosine_score"], reverse=True)
    if reset:
        pipe.reset_ledger()
    misses_before = pipe.ledger.misses
    n_reranked, cap_hit = 0, False
    for r in order:
        if max_rerank is not None and n_reranked >= max_rerank:
            cap_hit = True
            break
        # Budget is charged on NEW (network) calls only; cached calls are free.
        if max_rerank is None and pipe.ledger.misses + top_k > max_calls:
            cap_hit = True
            break
        res = pipe.classify(r["utterance"], top_k=top_k)
        top = res["ranked_candidates"][0] if res["ranked_candidates"] else {}
        r["llm_score"] = res["score"]
        r["degraded"] = res["degraded"]
        r["scored_by"] = "llm"
        r["rationale"] = top.get("rationale")
        r["top_match_in_corpus"] = top.get("utterance", r["top_match_in_corpus"])
        r["top_match_score"] = (top.get("llm_score")
                                if top.get("llm_score") is not None
                                else r["top_match_score"])
        n_reranked += 1
        if verbose and n_reranked % 25 == 0:
            print(f"    re-ranked {n_reranked} (new calls={pipe.ledger.misses})")

    folder_new_calls = pipe.ledger.misses - misses_before
    return {"n_reranked": n_reranked,
            "new_calls": folder_new_calls,                 # this call only
            "cumulative_new_calls": pipe.ledger.misses,    # whole run so far
            "cached_calls": pipe.ledger.hits,
            "dollars_spent": round(folder_new_calls * cost, 4),
            "cumulative_dollars": round(pipe.ledger.misses * cost, 4),
            "cap_hit": cap_hit, "max_calls": max_calls}


def assign_label(rec: dict, cfg: dict) -> None:
    """Predicted label + a unified display `score` (in place)."""
    if rec.get("scored_by") == "llm" and rec.get("llm_score") is not None:
        score = rec["llm_score"]
        rec["predicted_label"] = (LABEL_SCIENCE
                                  if score >= cfg.get("llm_label_threshold", 0.9)
                                  else LABEL_NOT)
    else:
        score = rec["cosine_score"]
        rec["predicted_label"] = (LABEL_SCIENCE
                                  if score >= cfg.get("cosine_label_threshold", 0.4)
                                  else LABEL_NOT)
    rec["score"] = float(score)


# ---------------------------------------------------------------------------
# Output workbook
# ---------------------------------------------------------------------------

_SHEET_COLS = [
    "line_no", "seg_start", "seg_end", "speaker", "child_id", "utterance",
    "confidence", "predicted_label", "review", "score", "scored_by",
    "cosine_score", "llm_score", "top_match_in_corpus", "top_match_score",
    "predicted_subtype", "rationale", "degraded",
]

def _build_instructions(science_only: bool) -> list[list[str]]:
    lines = [
        ["Y2 science-talk scoring — review guide"],
        [""],
        ["Each sheet is one child's point-of-view recorder for this classroom session."],
    ]
    if science_only:
        lines.append(["Sheets contain ONLY utterances the model labeled SCIENCE_TALK "
                      "(non-science rows were filtered out)."])
    else:
        lines.append(["One row per spoken utterance (kept rows: adult/child speech "
                      "above the confidence floor)."])
    lines += [
        [""],
        ["Columns:"],
        ["  predicted_label  : SCIENCE_TALK / NOT_SCIENCE_TALK (PROVISIONAL Y1-tuned threshold)."],
        ["  review           : YOU fill this in -- pick SCIENCE_TALK or NOT_SCIENCE_TALK from the dropdown."],
        ["  score            : the science-talk confidence used for the label (0-1)."],
        ["  scored_by        : 'llm' = refined by gpt-oss-120b; 'cosine' = bi-encoder only."],
        ["  cosine_score     : bi-encoder similarity to the nearest known science utterance."],
        ["  llm_score        : gpt-oss-120b judgment (only for the top science candidates)."],
        ["  top_match_in_corpus : the most similar known science utterance."],
        ["  rationale        : the model's short justification (llm rows only)."],
        [""],
        ["How to review:"],
        ["  1. Sort a sheet by 'score' descending to see the strongest science talk first."],
        ["  2. In the 'review' column, record your own judgment for each utterance."],
        ["  3. The tally at the bottom of each sheet counts how many of your reviews"],
        ["     AGREE with the model's predicted_label (and the agreement rate)."],
    ]
    if not science_only:
        lines.append(["  4. The SUMMARY sheet lists the top science utterances across all children."])
    lines += [
        ["  5. Labels are hints; the score + rationale are the evidence."],
        ["  6. Spot-check some high AND low scored rows and note the false-positive rate."],
    ]
    return lines


def _add_review_tools(ws, *, n_rows: int) -> None:
    """Add a SCIENCE/NOT dropdown to the `review` column and an agreement tally
    at the bottom of a per-child sheet. `n_rows` = number of data rows."""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    if n_rows <= 0:
        return
    rev_i = _SHEET_COLS.index("review") + 1
    pred_i = _SHEET_COLS.index("predicted_label") + 1
    rev_c = get_column_letter(rev_i)
    pred_c = get_column_letter(pred_i)
    first, last = 2, n_rows + 1  # data rows (row 1 is the header)
    rev_rng = f"{rev_c}{first}:{rev_c}{last}"
    pred_rng = f"{pred_c}{first}:{pred_c}{last}"

    dv = DataValidation(
        type="list",
        formula1=f'"{LABEL_SCIENCE},{LABEL_NOT}"',
        allow_blank=True,
        showDropDown=False,  # False => the dropdown arrow IS shown (openpyxl quirk)
    )
    dv.prompt = "Enter your judgment for this utterance."
    dv.promptTitle = "Reviewer label"
    ws.add_data_validation(dv)
    dv.add(rev_rng)

    s = last + 2  # leave a blank row, then the tally block
    ws.cell(row=s, column=1, value="Reviewed (non-blank):")
    ws.cell(row=s, column=2, value=f'=COUNTIF({rev_rng},"<>")')
    ws.cell(row=s + 1, column=1, value="Agreements with model:")
    ws.cell(row=s + 1, column=2, value=f"=SUMPRODUCT(--({rev_rng}={pred_rng}))")
    ws.cell(row=s + 2, column=1, value="Agreement rate:")
    ws.cell(row=s + 2, column=2, value=f'=IFERROR(B{s + 1}/B{s},"")')


def write_classroom_workbook(folder_name: str, recs: list[dict], cfg: dict,
                             *, verbose: bool = True) -> Path | None:
    science_only = cfg.get("science_only", True)
    if science_only:
        recs = [r for r in recs if r.get("predicted_label") == LABEL_SCIENCE]
    if not recs:
        if verbose:
            print(f"  skip {folder_name}: no SCIENCE_TALK utterances to write")
        return None

    out_dir = _PROJECT_ROOT / cfg.get("output_dir", "reports/y2_scores")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{folder_name}.xlsx"

    by_file: dict[str, list[dict]] = {}
    for r in recs:
        by_file.setdefault(r["source_file"], []).append(r)

    used: set[str] = set()
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        # SUMMARY (full-transcript mode only; science-only sheets are already science)
        if not science_only:
            top_n = cfg.get("summary_top_n", 50)
            ssorted = sorted(recs, key=lambda r: r["score"], reverse=True)[:top_n]
            summary = pd.DataFrame([{
                "score": r["score"], "scored_by": r["scored_by"],
                "child_id": r["child_id"], "source_file": r["source_file"],
                "line_no": r["line_no"], "speaker": r["speaker"],
                "utterance": r["utterance"], "predicted_label": r["predicted_label"],
                "predicted_subtype": r["predicted_subtype"],
                "top_match_in_corpus": r["top_match_in_corpus"],
                "rationale": r["rationale"],
            } for r in ssorted])
            (summary if not summary.empty
             else pd.DataFrame(columns=["score"])).to_excel(
                xl, sheet_name="SUMMARY", index=False)

        # one sheet per child file, chronological
        for fname in sorted(by_file):
            rows = sorted(by_file[fname], key=lambda r: r["line_no"])
            child = rows[0]["child_id"] if rows else "base"
            sn = sheet_name_for(child, used)
            df = pd.DataFrame(rows)
            for c in _SHEET_COLS:
                if c not in df.columns:
                    df[c] = None
            df["review"] = None  # reviewer fills this in
            df[_SHEET_COLS].to_excel(xl, sheet_name=sn, index=False)
            _add_review_tools(xl.sheets[sn], n_rows=len(df))

        pd.DataFrame(_build_instructions(science_only)).to_excel(
            xl, sheet_name="INSTRUCTIONS", index=False, header=False)

    if verbose:
        kind = "science " if science_only else ""
        print(f"  wrote {out_path.relative_to(_PROJECT_ROOT)} "
              f"({len(by_file)} child sheets, {len(recs)} {kind}utterances)")
    return out_path


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def _resolve_folders(cfg: dict, folders: list[str] | None, all_: bool) -> list[Path]:
    root = _PROJECT_ROOT / cfg["transcripts_root"]
    if folders:
        return [root / f for f in folders]
    if not all_:
        raise RuntimeError("Specify --folders <name...> for a pilot, or --all.")
    keep_suffix = cfg.get("classroom49_keep_suffix", "_updated")
    out = []
    for d in sorted(p for p in root.iterdir() if p.is_dir()):
        if d.name.startswith("Classroom49_012026"):
            if not d.name.endswith(keep_suffix):
                continue
        out.append(d)
    return out


def run(
    folders: list[str] | None = None,
    *,
    config_path: Path = DEFAULT_CONFIG_PATH,
    all_: bool = False,
    use_stub: bool = False,
    processed_dir: Path = DEFAULT_PROCESSED_DIR,
    overrides: dict | None = None,
    overwrite: bool = False,
    verbose: bool = True,
) -> dict:
    """Score Y2 transcripts folder-by-folder, writing each workbook as it
    finishes (memory-bounded + resumable). The LLM budget is enforced on
    cumulative spend across folders via a single persistent ledger; finished
    folders are skipped on a re-run unless `overwrite=True`.
    """
    from src.query_10 import QueryPipeline
    from src import reranker_9 as r9

    cfg = load_config(config_path)
    cfg.update(overrides or {})
    targets = _resolve_folders(cfg, folders, all_)
    out_dir = _PROJECT_ROOT / cfg.get("output_dir", "reports/y2_scores")
    out_dir.mkdir(parents=True, exist_ok=True)
    cost = cfg.get("cost_per_call", 0.0001)
    dollar_cap = cfg.get("dollar_cap", 15.0)
    share = dollar_cap / max(len(targets), 1)  # per-folder allowance (rolls over)
    if verbose:
        print(f"Step 12 deploy | folders={len(targets)} | stub={use_stub} | "
              f"cap=${dollar_cap} top_k={cfg['top_k']} | resume={not overwrite}")

    pipe = QueryPipeline(
        processed_dir, config_path=_PROJECT_ROOT / cfg["query_config"],
        caller=(r9.stub_caller if use_stub else None), verbose=False)
    pipe.reset_ledger()  # one ledger for the whole run (cumulative cap)

    per_folder_report: list[dict] = []
    out_paths: list[str] = []
    totals = {"n_utterances": 0, "n_science": 0, "done": 0, "skipped": 0}

    def _write_manifest() -> dict:
        manifest = {
            "version": cfg["version"], "stub": use_stub,
            "dollar_cap": dollar_cap, "top_k": cfg["top_k"],
            "science_only": cfg.get("science_only", True),
            "n_folders_selected": len(targets),
            "n_folders_done": totals["done"], "n_folders_skipped": totals["skipped"],
            "n_utterances": totals["n_utterances"],
            "n_predicted_science": totals["n_science"],
            "cumulative_new_calls": pipe.ledger.misses,
            "cumulative_dollars": round(pipe.ledger.misses * cost, 4),
            "cached_calls": pipe.ledger.hits,
            "cap_reached": pipe.ledger.misses * cost >= dollar_cap - cost,
            "per_folder": per_folder_report,
            "workbooks": out_paths,
        }
        with open(out_dir / "run_manifest.json", "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2, default=str)
        return manifest

    for k, folder in enumerate(targets):
        if not folder.exists():
            print(f"  [warn] missing folder {folder.name}")
            continue
        out_path = out_dir / f"{folder.name}.xlsx"
        if out_path.exists() and not overwrite:
            if verbose:
                print(f"  skip {folder.name}: workbook already exists (resume).")
            totals["skipped"] += 1
            out_paths.append(str(out_path))
            continue

        recs: list[dict] = []
        for xlsx in sorted(folder.glob("*.xlsx")):
            recs.extend(ingest_file(xlsx, cfg))
        if not recs:
            if verbose:
                print(f"  {folder.name}: no utterances after filters; skipping.")
            continue
        if verbose:
            print(f"  [{k + 1}/{len(targets)}] {folder.name}: {len(recs)} utterances")

        cosine_score_all(pipe, recs, verbose=verbose)

        if use_stub:
            led = budgeted_rerank(pipe, recs, cfg, max_rerank=30, reset=False,
                                  verbose=verbose)
        else:
            # cumulative allowance grows by one share per folder, capped at the
            # total; unused budget from quiet folders rolls forward.
            allowance = min(dollar_cap, (k + 1) * share)
            led = budgeted_rerank(pipe, recs, cfg, dollar_budget=allowance,
                                  reset=False, verbose=verbose)

        for r in recs:
            assign_label(r, cfg)

        path = write_classroom_workbook(folder.name, recs, cfg, verbose=verbose)
        if path is not None:
            out_paths.append(str(path))

        n_sci = sum(1 for r in recs if r["predicted_label"] == LABEL_SCIENCE)
        totals["n_utterances"] += len(recs)
        totals["n_science"] += n_sci
        totals["done"] += 1
        per_folder_report.append({
            "folder": folder.name, "utterances": len(recs),
            "predicted_science": n_sci, "reranked": led["n_reranked"],
            "new_calls": led["new_calls"], "dollars_this_folder": led["dollars_spent"],
            "cumulative_dollars": led["cumulative_dollars"],
            "workbook": str(path) if path else None,
        })
        _write_manifest()  # checkpoint after every folder
        del recs  # free memory before the next folder

        if pipe.ledger.misses * cost >= dollar_cap - cost:
            if verbose:
                print(f"  budget cap ${dollar_cap} reached; remaining folders "
                      f"will be cosine-only.")

    manifest = _write_manifest()
    if verbose:
        _print_summary(manifest)
    return manifest


def _print_summary(m: dict) -> None:
    print("\n=== Step 12 Y2 scoring ===")
    print(f"  folders done: {m['n_folders_done']} | skipped: "
          f"{m['n_folders_skipped']} / selected {m['n_folders_selected']}")
    print(f"  utterances: {m['n_utterances']} | predicted science: "
          f"{m['n_predicted_science']} "
          f"({m['n_predicted_science'] / max(m['n_utterances'], 1):.1%})")
    print(f"  LLM spend: {m['cumulative_new_calls']} new calls | "
          f"${m['cumulative_dollars']} (cap ${m['dollar_cap']}, "
          f"reached={m['cap_reached']})")
    print(f"  workbooks: {len(m['workbooks'])}")


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawTextHelpFormatter)
    p.add_argument("--folders", nargs="+", default=None,
                   help="Classroom-session subfolder name(s) for a pilot.")
    p.add_argument("--all", action="store_true", help="Process all 85 folders.")
    p.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    p.add_argument("--stub", action="store_true", help="Offline (cosine real, no LLM).")
    p.add_argument("--dollar-cap", type=float, default=None)
    p.add_argument("--top-k", type=int, default=None)
    p.add_argument("--overwrite", action="store_true",
                   help="Re-score folders even if their workbook already exists "
                        "(default: skip finished folders so a run resumes).")
    p.add_argument("--quiet", action="store_true")
    return p.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    ov = {}
    if args.dollar_cap is not None:
        ov["dollar_cap"] = args.dollar_cap
    if args.top_k is not None:
        ov["top_k"] = args.top_k
    run(args.folders, config_path=args.config, all_=args.all, overwrite=args.overwrite,
        use_stub=args.stub, overrides=ov, verbose=not args.quiet)
