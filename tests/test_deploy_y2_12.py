"""Unit tests for src/deploy_y2_12.py (Step 12 Y2 deployment scoring).

Offline only: a tiny fake bi-encoder + a fake pipeline stand in for the real
model and the LLM endpoint. The real pilot is run via the CLI.
"""

from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from src import deploy_y2_12 as d12


# ---------------------------------------------------------------------------
# parse_transcription
# ---------------------------------------------------------------------------

def test_parse_transcription_with_confidence():
    text, conf = d12.parse_transcription("it's reading 16. (57.6%)")
    assert text == "it's reading 16." and conf == pytest.approx(0.576)


def test_parse_transcription_integer_pct():
    assert d12.parse_transcription("No. (76%)") == ("No.", 0.76)


def test_parse_transcription_no_confidence():
    assert d12.parse_transcription("hello there") == ("hello there", None)


def test_parse_transcription_blank():
    assert d12.parse_transcription(None) == ("", None)


# ---------------------------------------------------------------------------
# filename / sheet helpers
# ---------------------------------------------------------------------------

def test_child_id_from_filename():
    assert d12.child_id_from_filename("65-74-943-170_110525_sync.xlsx") == "170"
    assert d12.child_id_from_filename("65-74-943_110525_sync.xlsx") == "base"


def test_sheet_name_unique_and_safe():
    used = set()
    a = d12.sheet_name_for("170", used)
    b = d12.sheet_name_for("170", used)  # collision -> suffixed
    assert a == "child_170" and b != a and len(b) <= 31


# ---------------------------------------------------------------------------
# ingest_file
# ---------------------------------------------------------------------------

def test_ingest_file_filters(tmp_path):
    df = pd.DataFrame({
        "Segment Start Time": ["00:00:01", "00:00:02", "-", "00:00:04", "00:00:05"],
        "Segment End Time": ["00:00:02", "00:00:03", "-", "00:00:05", "00:00:06"],
        "Transcription (Confidence)": [
            "what do you notice? (90%)",   # adult, high conf -> keep
            "i think it floats (30%)",      # child, low conf -> drop
            "-",                             # proximity row -> drop
            "clean up please (88%)",        # unknown speaker -> drop
            "the ice melted (85%)",          # child, ok -> keep
        ],
        "Speaker": ["adult", "child", "-", "unknown", "child"],
    })
    f = tmp_path / "65-74-943-170_110525_sync.xlsx"
    df.to_excel(f, index=False)

    cfg = {"speaker_keep": ["adult", "child"], "confidence_min": 0.40}
    recs = d12.ingest_file(f, cfg)
    texts = [r["utterance"] for r in recs]
    assert texts == ["what do you notice?", "the ice melted"]
    assert all(r["child_id"] == "170" for r in recs)
    assert recs[0]["speaker"] == "adult" and recs[0]["confidence"] == 0.90


# ---------------------------------------------------------------------------
# _primary_subtype
# ---------------------------------------------------------------------------

def test_primary_subtype():
    assert d12._primary_subtype(["observation"]) == "observation"
    assert d12._primary_subtype(["b", "a"]) == "a|b"
    assert d12._primary_subtype(None) == "unknown"


# ---------------------------------------------------------------------------
# Fakes for cosine + rerank
# ---------------------------------------------------------------------------

class FakeST:
    def get_sentence_embedding_dimension(self):
        return 3

    def encode(self, texts, **kwargs):
        out = []
        for t in texts:
            tl = t.lower()
            v = [1.0, 0.0, 0.0] if "ice" in tl else \
                [0.0, 1.0, 0.0] if "plant" in tl else [0.0, 0.0, 1.0]
            out.append(v)
        a = np.asarray(out, dtype=np.float32)
        n = np.linalg.norm(a, axis=1, keepdims=True)
        n[n == 0] = 1.0
        return a / n


class FakePipe:
    def __init__(self):
        import src.reranker_9 as r9
        self.model = FakeST()
        self.corpus_emb = np.eye(3, dtype=np.float32)
        self.meta_utt_ids = ["ice_a", "plant_a", "other_a"]
        self.meta_utts = ["why did the ice melt", "how did the plant grow", "blocks"]
        self.subtype_map = {"ice_a": ["observation"], "plant_a": ["content"]}
        self.query_prefix = ""
        self.ledger = r9.CallLedger()

    def reset_ledger(self):
        self.ledger.hits = 0
        self.ledger.misses = 0

    def classify(self, phrase, *, top_k=10):
        # pretend each classify makes top_k new endpoint calls
        self.ledger.misses += top_k
        sci = "ice" in phrase.lower() or "plant" in phrase.lower()
        score = 0.95 if sci else 0.05
        return {"score": score, "degraded": False,
                "ranked_candidates": [{"utt_id": "ice_a", "utterance": "why did the ice melt",
                                       "llm_score": score, "bi_score": 0.5,
                                       "rationale": "sci" if sci else "no"}]}


def _records(texts):
    return [{"child_id": "170", "source_file": "f.xlsx", "line_no": i + 2,
             "seg_start": "", "seg_end": "", "speaker": "child", "utterance": t,
             "confidence": 0.9} for i, t in enumerate(texts)]


# ---------------------------------------------------------------------------
# cosine_score_all
# ---------------------------------------------------------------------------

def test_cosine_score_all():
    pipe = FakePipe()
    recs = _records(["the ice is cold", "a plant grows", "random chatter"])
    d12.cosine_score_all(pipe, recs, verbose=False)
    assert recs[0]["top_match_utt_id"] == "ice_a"
    assert recs[0]["cosine_score"] == pytest.approx(1.0)
    assert recs[0]["predicted_subtype"] == "observation"
    assert all(r["scored_by"] == "cosine" for r in recs)


# ---------------------------------------------------------------------------
# budgeted_rerank
# ---------------------------------------------------------------------------

def test_budgeted_rerank_dollar_cap():
    pipe = FakePipe()
    recs = _records(["ice one", "ice two", "plant three", "blah four", "blah five"])
    d12.cosine_score_all(pipe, recs, verbose=False)
    # cap = 50 calls; top_k=10 -> exactly 5 reranks possible, but only science
    # rank high; still limited by dollars
    cfg = {"top_k": 10, "dollar_cap": 0.003, "cost_per_call": 0.0001}  # 30 calls -> 3 reranks
    led = d12.budgeted_rerank(pipe, recs, cfg, verbose=False)
    assert led["n_reranked"] == 3
    assert led["new_calls"] == 30
    assert led["dollars_spent"] == pytest.approx(0.003)
    assert led["cap_hit"] is True


def test_budgeted_rerank_max_rerank_cap():
    pipe = FakePipe()
    recs = _records(["ice a", "plant b", "c", "d"])
    d12.cosine_score_all(pipe, recs, verbose=False)
    led = d12.budgeted_rerank(pipe, recs, {"top_k": 10, "dollar_cap": 99,
                                           "cost_per_call": 0.0001},
                              max_rerank=2, verbose=False)
    assert led["n_reranked"] == 2
    reranked = [r for r in recs if r["scored_by"] == "llm"]
    assert len(reranked) == 2


# ---------------------------------------------------------------------------
# assign_label
# ---------------------------------------------------------------------------

def test_assign_label_llm_and_cosine():
    cfg = {"llm_label_threshold": 0.9, "cosine_label_threshold": 0.4}
    llm = {"scored_by": "llm", "llm_score": 0.95, "cosine_score": 0.1}
    d12.assign_label(llm, cfg)
    assert llm["predicted_label"] == d12.LABEL_SCIENCE and llm["score"] == 0.95

    cos = {"scored_by": "cosine", "llm_score": None, "cosine_score": 0.3}
    d12.assign_label(cos, cfg)
    assert cos["predicted_label"] == d12.LABEL_NOT and cos["score"] == 0.3


# ---------------------------------------------------------------------------
# write_classroom_workbook
# ---------------------------------------------------------------------------

def _scored_record(text, child, fname, line, label, *, score=0.5):
    return {
        "child_id": child, "source_file": fname, "line_no": line,
        "seg_start": "", "seg_end": "", "speaker": "child", "utterance": text,
        "confidence": 0.9, "cosine_score": score, "llm_score": None,
        "top_match_in_corpus": "anchor", "top_match_score": score,
        "predicted_subtype": "observation", "rationale": None, "degraded": None,
        "scored_by": "cosine", "predicted_label": label, "score": score,
    }


def _data_rows(xl, sheet):
    """Real utterance rows only (excludes the appended tally block)."""
    df = xl.parse(sheet)
    return df[df["utterance"].notna()]


def test_write_classroom_workbook_full_mode(tmp_path):
    pipe = FakePipe()
    recs = _records(["the ice is cold", "random chatter"])
    recs[1]["child_id"] = "511"
    recs[1]["source_file"] = "g.xlsx"
    d12.cosine_score_all(pipe, recs, verbose=False)
    cfg = {"output_dir": str(tmp_path), "summary_top_n": 10, "science_only": False,
           "llm_label_threshold": 0.9, "cosine_label_threshold": 0.4}
    for r in recs:
        d12.assign_label(r, cfg)
    out = d12.write_classroom_workbook("Classroom74_110525", recs, cfg, verbose=False)
    assert out.exists()
    xl = pd.ExcelFile(out)
    assert "SUMMARY" in xl.sheet_names and "INSTRUCTIONS" in xl.sheet_names
    child_sheets = [s for s in xl.sheet_names if s.startswith("child_")]
    assert len(child_sheets) == 2
    df = xl.parse(child_sheets[0])
    assert "predicted_label" in df.columns and "score" in df.columns
    assert "review" in df.columns


def test_write_classroom_workbook_science_only(tmp_path):
    # child 170: one science + one non-science; child 511: all non-science
    recs = [
        _scored_record("the ice is cold", "170", "f.xlsx", 2, d12.LABEL_SCIENCE),
        _scored_record("um okay", "170", "f.xlsx", 3, d12.LABEL_NOT),
        _scored_record("random chatter", "511", "g.xlsx", 2, d12.LABEL_NOT),
    ]
    cfg = {"output_dir": str(tmp_path), "science_only": True}
    out = d12.write_classroom_workbook("Classroom74_110525", recs, cfg, verbose=False)
    xl = pd.ExcelFile(out)
    # SUMMARY omitted in science-only mode
    assert "SUMMARY" not in xl.sheet_names and "INSTRUCTIONS" in xl.sheet_names
    # child 511 dropped (no science); child 170 keeps only its science row
    child_sheets = [s for s in xl.sheet_names if s.startswith("child_")]
    assert len(child_sheets) == 1
    data = _data_rows(xl, child_sheets[0])
    assert len(data) == 1
    assert set(data["predicted_label"]) == {d12.LABEL_SCIENCE}
    # empty review column present
    assert "review" in data.columns and data["review"].isna().all()


def test_review_column_and_agreement_tally(tmp_path):
    import openpyxl
    recs = [
        _scored_record("the ice is cold", "170", "f.xlsx", 2, d12.LABEL_SCIENCE),
        _scored_record("water freezes", "170", "f.xlsx", 3, d12.LABEL_SCIENCE),
    ]
    cfg = {"output_dir": str(tmp_path), "science_only": True}
    out = d12.write_classroom_workbook("Classroom74_110525", recs, cfg, verbose=False)
    wb = openpyxl.load_workbook(out)  # formulas (not evaluated)
    ws = wb[[s for s in wb.sheetnames if s.startswith("child_")][0]]
    header = [c.value for c in ws[1]]
    assert "review" in header
    # locate the tally block by its labels in column A
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    assert "Agreements with model:" in labels
    agree_row = labels["Agreements with model:"]
    assert str(ws.cell(row=agree_row, column=2).value).startswith("=SUMPRODUCT")
    assert "Reviewed (non-blank):" in labels and "Agreement rate:" in labels
    # a data-validation dropdown is attached to the review column
    assert len(ws.data_validations.dataValidation) >= 1


def test_write_classroom_workbook_no_science_returns_none(tmp_path):
    recs = [_scored_record("um okay", "170", "f.xlsx", 2, d12.LABEL_NOT)]
    cfg = {"output_dir": str(tmp_path), "science_only": True}
    out = d12.write_classroom_workbook("Classroom1_x", recs, cfg, verbose=False)
    assert out is None


def test_config_loads():
    cfg = d12.load_config()
    assert cfg["top_k"] == 10 and cfg["dollar_cap"] == 15.0
    assert cfg["speaker_keep"] == ["adult", "child"]
    assert cfg["science_only"] is True
